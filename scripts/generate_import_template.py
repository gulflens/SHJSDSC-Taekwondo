"""
Generate templates/SHJSDSC_import_template.xlsx — a single Excel workbook with
one sheet per importable entity (Branches, Coaches, Athletes), plus a README
and an Enums reference. Column names match the keys the in-app CSV importer
already understands so a row authored here can be saved as CSV (Athletes only,
for now) and imported via Athletes → toolbar → Import.

Re-run after model changes: `python3 scripts/generate_import_template.py`.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent.parent / "templates" / "SHJSDSC_import_template.xlsx"

REQUIRED_FILL = PatternFill("solid", fgColor="FFE7B0")   # warm amber for required
HEADER_FILL = PatternFill("solid", fgColor="1F3B70")     # SHJSDSC navy
HINT_FILL = PatternFill("solid", fgColor="F2F4F8")       # neutral gray
EXAMPLE_FILL = PatternFill("solid", fgColor="E8F2EA")    # green-tinted example row

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HINT_FONT = Font(italic=True, color="555555", size=10)
EXAMPLE_FONT = Font(color="2D5A33", size=10)

THIN = Side(border_style="thin", color="BFC7D2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ----- columns -----
# Each column: (key, friendly_label, required, hint, example, [enum_values])

BRANCH_COLUMNS = [
    ("code", "Branch code", True, "Unique short code, e.g. BR-RAH", "BR-RAH", None),
    ("name", "Name (English)", True, "Public-facing English name", "Al Rahmania", None),
    ("name_ar", "Name (Arabic)", True, "Arabic name", "الرحمانية", None),
    ("area", "Area / district", True, "Neighbourhood within emirate", "Al Rahmania", None),
    ("capacity", "Capacity (max athletes)", True, "Integer ≥ 0", 120, None),
    ("focus", "Focus", True, "Free-text: poomsae / kyorugi / mixed", "mixed", None),
    ("manager_email", "Manager email", False, "Email of the user_profile to set as manager (optional)", "", None),
    ("street_address", "Street address (English)", False, "", "Rahmania Street 12", None),
    ("street_address_ar", "Street address (Arabic)", False, "", "شارع الرحمانية 12", None),
    ("emirate", "Emirate", False, "Default Sharjah", "Sharjah", None),
    ("country", "Country (ISO-2)", False, "Default AE", "AE", None),
    ("po_box", "PO Box", False, "", "12345", None),
    ("latitude", "Latitude", False, "Decimal degrees", 25.3463, None),
    ("longitude", "Longitude", False, "Decimal degrees", 55.4209, None),
    ("google_place_id", "Google Place ID", False, "From Google Maps", "", None),
    ("phone", "Phone", False, "Local format", "+971-6-555-1234", None),
    ("whatsapp_business", "WhatsApp business", False, "International format", "+971501234567", None),
    ("email", "Email", False, "", "info@example.ae", None),
    ("founded_at", "Founded date", False, "YYYY-MM-DD", "2018-09-01", None),
    ("is_active", "Is active", False, "TRUE / FALSE (default TRUE)", "TRUE", ["TRUE", "FALSE"]),
    ("brand_hex_color", "Brand colour (hex)", False, "e.g. #1F3B70", "#1F3B70", None),
    ("tagline_en", "Tagline (English)", False, "", "Excellence in motion", None),
    ("tagline_ar", "Tagline (Arabic)", False, "", "التميز في الحركة", None),
]

COACH_COLUMNS = [
    ("full_name", "Full name (English)", True, "", "Yassin Al-Jawadi", None),
    ("full_name_ar", "Full name (Arabic)", True, "", "ياسين الجوادي", None),
    ("primary_branch_code", "Primary branch code", True, "Must match a code on the Branches sheet", "BR-NAS", None),
    ("secondary_branch_codes", "Secondary branch codes", False, "Semicolon-separated list", "BR-RAH;BR-NOUF", None),
    ("dan_rank", "Dan rank", True, "Integer 0–9 (0 if pre-black)", 4, None),
    ("wt_coach_licence_level", "WT coach licence level", True, "Integer 0–3", 2, None),
    ("first_aid_expiry", "First-aid cert expiry", True, "YYYY-MM-DD", "2027-04-30", None),
    ("safeguarding_expiry", "Safeguarding cert expiry", True, "YYYY-MM-DD", "2027-04-30", None),
    ("contract_type", "Contract type", True, "fullTime / partTime / contractor", "fullTime", ["fullTime", "partTime", "contractor"]),
    ("hired_at", "Hired date", True, "YYYY-MM-DD", "2020-09-01", None),
    ("kukkiwon_cert_number", "Kukkiwon cert #", False, "", "K-12345", None),
    ("kukkiwon_issued_at", "Kukkiwon issue date", False, "YYYY-MM-DD", "2018-06-01", None),
    ("wt_coach_licence_expiry", "WT licence expiry", False, "YYYY-MM-DD", "2026-12-31", None),
    ("poomsae_referee_level", "Poomsae referee level", False, "Integer 0–4", 2, None),
    ("poomsae_referee_expiry", "Poomsae referee expiry", False, "YYYY-MM-DD", "2026-09-30", None),
    ("kyorugi_referee_level", "Kyorugi referee level", False, "Integer 0–4", 1, None),
    ("kyorugi_referee_expiry", "Kyorugi referee expiry", False, "YYYY-MM-DD", "2026-09-30", None),
    ("anti_doping_expiry", "Anti-doping cert expiry", False, "YYYY-MM-DD", "2027-01-31", None),
    ("weekly_hours_target", "Weekly hours target", False, "Integer", 30, None),
    ("on_call", "On-call", False, "TRUE / FALSE", "FALSE", ["TRUE", "FALSE"]),
    ("bio", "Bio (English)", False, "Short paragraph", "Former national team athlete; coaches juniors.", None),
    ("bio_ar", "Bio (Arabic)", False, "", "لاعب سابق في المنتخب الوطني، يدرب الناشئين.", None),
    ("avatar_url", "Avatar URL", False, "Public image URL", "", None),
]

ATHLETE_COLUMNS = [
    ("full_name", "Full name (English)", True, "", "Ahmed Al Mazrouei", None),
    ("full_name_ar", "Full name (Arabic)", True, "", "أحمد المزروعي", None),
    ("date_of_birth", "Date of birth", True, "YYYY-MM-DD", "2012-03-14", None),
    ("gender", "Gender", True, "male / female", "male", ["male", "female"]),
    ("branch_code", "Branch code", True, "Must match a code on the Branches sheet", "BR-NAS", None),
    ("joined_at", "Joined date", True, "YYYY-MM-DD", "2022-09-01", None),
    ("weight_kg", "Weight (kg)", True, "Decimal", 48, None),
    ("current_belt_color", "Belt colour", True, "white / yellow / green / blue / red / black",
        "blue", ["white", "yellow", "green", "blue", "red", "black"]),
    ("current_belt_kind", "Belt kind", True, "gup / poom / dan", "gup", ["gup", "poom", "dan"]),
    ("current_belt_number", "Belt number", True, "Integer (10=white, 1=red, dan 1–9)", 4, None),
    ("status", "Athlete status", True, "active / competitionTeam / readyToGrade / watch / rest",
        "active", ["active", "competitionTeam", "readyToGrade", "watch", "rest"]),
    ("nationality", "Nationality (ISO-2)", False, "Default AE", "AE", None),
    ("emirates_id", "Emirates ID", False, "Format 784-YYYY-NNNNNNN-N", "784-1985-1234567-8", None),
    ("passport_number", "Passport #", False, "", "P12345678", None),
    ("blood_type", "Blood type", False, "A+ A- B+ B- AB+ AB- O+ O-",
        "A+", ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]),
    ("federation_licence_number", "Federation licence #", False, "", "FED-12345", None),
    ("school", "School", False, "", "Sharjah Public School", None),
    ("height_cm", "Height (cm)", False, "Decimal", 148, None),
    ("allergies", "Allergies", False, "Semicolon-separated", "peanuts;dairy", None),
    ("medical_conditions", "Medical conditions", False, "Semicolon-separated", "asthma", None),
    ("medications", "Medications", False, "Semicolon-separated", "inhaler", None),
    ("fit_to_train", "Fit to train", False, "TRUE / FALSE (default TRUE)", "TRUE", ["TRUE", "FALSE"]),
    ("dominant_stance", "Dominant stance", False, "open / closed / switchStance",
        "open", ["open", "closed", "switchStance"]),
    ("poomsae_syllabus", "Poomsae syllabus", False, "e.g. taegeuk-7", "taegeuk-7", None),
    ("kyorugi_tier", "Kyorugi tier", False, "recreational / competitive / elite",
        "competitive", ["recreational", "competitive", "elite"]),
    ("primary_coach_full_name", "Primary coach full name", False,
        "Must match an existing coach's full_name exactly", "Yassin Al-Jawadi", None),
]


def write_entity_sheet(wb, title, columns, second_example=None):
    ws = wb.create_sheet(title)

    # Row 1: technical column key (this is what the importer reads).
    # Row 2: friendly label.
    # Row 3: hint / format note.
    # Row 4+: example rows.
    for idx, (key, label, required, hint, example, _enum) in enumerate(columns, start=1):
        col = get_column_letter(idx)

        header = ws.cell(row=1, column=idx, value=key)
        header.fill = HEADER_FILL
        header.font = HEADER_FONT
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header.border = BORDER

        label_cell = ws.cell(row=2, column=idx, value=("* " if required else "") + label)
        label_cell.fill = REQUIRED_FILL if required else HINT_FILL
        label_cell.font = Font(bold=required, size=10)
        label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        label_cell.border = BORDER

        hint_cell = ws.cell(row=3, column=idx, value=hint)
        hint_cell.fill = HINT_FILL
        hint_cell.font = HINT_FONT
        hint_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hint_cell.border = BORDER

        example_cell = ws.cell(row=4, column=idx, value=example)
        example_cell.fill = EXAMPLE_FILL
        example_cell.font = EXAMPLE_FONT
        example_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        example_cell.border = BORDER

        if second_example is not None:
            second_cell = ws.cell(row=5, column=idx, value=second_example.get(key, ""))
            second_cell.fill = EXAMPLE_FILL
            second_cell.font = EXAMPLE_FONT
            second_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            second_cell.border = BORDER

        ws.column_dimensions[col].width = max(14, min(34, len(label) + 6))

    # Data validation for enum columns. Apply to a wide range so the dropdown
    # is available for every future entry the user types.
    for idx, (_key, _label, _req, _hint, _ex, enum) in enumerate(columns, start=1):
        if not enum:
            continue
        col = get_column_letter(idx)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(enum) + '"',
            allow_blank=True,
            showDropDown=False,  # openpyxl: False = arrow visible
        )
        dv.error = "Pick one of: " + ", ".join(enum)
        dv.errorTitle = "Invalid value"
        dv.add(f"{col}4:{col}1000")
        ws.add_data_validation(dv)

    # Freeze header rows so they stay visible while scrolling.
    ws.freeze_panes = "A4"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 38

    # Right-align Arabic columns visually (still LTR per CLAUDE.md).
    for idx, (key, *_rest) in enumerate(columns, start=1):
        if key.endswith("_ar"):
            for row in range(4, 1001):
                ws.cell(row=row, column=idx).alignment = Alignment(horizontal="right", vertical="center")


def write_readme(wb):
    ws = wb.create_sheet("Read me first", 0)

    blocks = [
        ("SHJSDSC Taekwondo — Bulk Import Template", 16, True, HEADER_FILL, "FFFFFF"),
        ("", 10, False, None, None),
        ("How the rows are arranged", 12, True, None, None),
        ("Row 1 (navy)  : technical column name — this is what the importer reads. Do not rename.", 10, False, None, None),
        ("Row 2 (amber) : friendly label. A leading * marks a required column.", 10, False, None, None),
        ("Row 3 (grey)  : format hint or list of accepted enum values.", 10, False, None, None),
        ("Row 4–5       : worked examples. Delete or overwrite when you are ready to enter your own data.", 10, False, None, None),
        ("Row 6+        : your data. One entity per row.", 10, False, None, None),
        ("", 10, False, None, None),
        ("Order to fill in", 12, True, None, None),
        ("1. Branches    — every Coach and Athlete references a branch_code, so create those first.", 10, False, None, None),
        ("2. Coaches     — each coach references primary_branch_code (must already exist on the Branches sheet).", 10, False, None, None),
        ("3. Athletes    — each athlete references branch_code and optionally primary_coach_full_name.", 10, False, None, None),
        ("", 10, False, None, None),
        ("Importing into the app today", 12, True, None, None),
        ("• Athletes: save the Athletes sheet as CSV (File → Save As → CSV UTF-8) and import via", 10, False, None, None),
        ("  Athletes tab → toolbar → Import (square.and.arrow.down).", 10, False, None, None),
        ("• Branches & Coaches: enter via Admin → Add Branch / Add Coach using the values below as your reference. ", 10, False, None, None),
        ("  A bulk importer for these two sheets is on the roadmap; the column keys here already match the model so it will read this file as-is.", 10, False, None, None),
        ("", 10, False, None, None),
        ("Conventions", 12, True, None, None),
        ("• Dates : YYYY-MM-DD (Excel date cells are also accepted; they're serialised the same way).", 10, False, None, None),
        ("• Lists : separate values with semicolons — e.g. allergies = peanuts;dairy.", 10, False, None, None),
        ("• Booleans : TRUE or FALSE (case-insensitive; yes/no/1/0/y/n also work).", 10, False, None, None),
        ("• Enums : pick from the dropdown; see the Enums sheet for the full reference.", 10, False, None, None),
        ("• Arabic text : paste directly. Cells are right-aligned but stored LTR — do not reverse character order.", 10, False, None, None),
        ("", 10, False, None, None),
        ("Server-allocated fields", 12, True, None, None),
        ("Do not include id, member_number, or avatar_seed — they are generated on save.", 10, False, None, None),
        ("", 10, False, None, None),
        ("On import errors", 12, True, None, None),
        ("Rows with an invalid required field, an unknown branch_code, or an unknown primary_coach_full_name", 10, False, None, None),
        ("are skipped with a row number + reason in the import summary. Fix the cell and re-import — the importer", 10, False, None, None),
        ("does NOT de-duplicate by name, so re-running creates new records. Edit existing entries in the app.", 10, False, None, None),
    ]

    for i, (text, size, bold, fill, font_color) in enumerate(blocks, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color=font_color or "1F2937")
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=False)

    ws.column_dimensions["A"].width = 130
    ws.row_dimensions[1].height = 30
    ws.sheet_view.showGridLines = False


def write_enums(wb):
    ws = wb.create_sheet("Enums")

    enums = [
        ("Athlete status", ["active", "competitionTeam", "readyToGrade", "watch", "rest"]),
        ("Athlete gender", ["male", "female"]),
        ("Belt colour", ["white", "yellow", "green", "blue", "red", "black"]),
        ("Belt kind", ["gup", "poom", "dan"]),
        ("Blood type", ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]),
        ("Dominant stance", ["open", "closed", "switchStance"]),
        ("Dominant leg", ["left", "right"]),
        ("Specialty", ["kyorugi", "poomsae", "both"]),
        ("Kyorugi tier", ["recreational", "competitive", "elite"]),
        ("Coach contract type", ["fullTime", "partTime", "contractor"]),
        ("Boolean", ["TRUE", "FALSE"]),
    ]

    ws.cell(row=1, column=1, value="Reference: valid values for enum-typed columns").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")

    for idx, (label, values) in enumerate(enums):
        col = get_column_letter(idx + 1)
        header = ws.cell(row=3, column=idx + 1, value=label)
        header.fill = HEADER_FILL
        header.font = HEADER_FONT
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header.border = BORDER
        for j, v in enumerate(values, start=4):
            cell = ws.cell(row=j, column=idx + 1, value=v)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
        ws.column_dimensions[col].width = max(14, len(label) + 4)

    ws.freeze_panes = "A4"
    ws.row_dimensions[3].height = 24
    ws.sheet_view.showGridLines = False


def main():
    wb = Workbook()
    # Drop the default "Sheet" — we add our own in order.
    wb.remove(wb.active)

    write_readme(wb)

    branch_second = {
        "code": "BR-NAS", "name": "Al Nasserya", "name_ar": "الناصرية",
        "area": "Al Nasserya", "capacity": 90, "focus": "kyorugi",
        "emirate": "Sharjah", "country": "AE", "is_active": "TRUE",
        "phone": "+971-6-555-7788", "email": "nas@example.ae",
        "founded_at": "2019-04-01",
    }
    write_entity_sheet(wb, "Branches", BRANCH_COLUMNS, second_example=branch_second)

    coach_second = {
        "full_name": "Dr Ali Alawi", "full_name_ar": "د. علي العلوي",
        "primary_branch_code": "BR-RAH", "dan_rank": 5,
        "wt_coach_licence_level": 3, "first_aid_expiry": "2027-02-15",
        "safeguarding_expiry": "2027-02-15", "contract_type": "partTime",
        "hired_at": "2017-09-01", "on_call": "TRUE", "weekly_hours_target": 12,
    }
    write_entity_sheet(wb, "Coaches", COACH_COLUMNS, second_example=coach_second)

    athlete_second = {
        "full_name": "Mariam Al Suwaidi", "full_name_ar": "مريم السويدي",
        "date_of_birth": "2015-07-22", "gender": "female",
        "branch_code": "BR-RAH", "joined_at": "2024-01-15", "weight_kg": 32,
        "current_belt_color": "yellow", "current_belt_kind": "gup",
        "current_belt_number": 8, "status": "active", "nationality": "AE",
        "blood_type": "O+", "school": "Al Rahmania School", "height_cm": 135,
        "fit_to_train": "TRUE", "dominant_stance": "open",
        "poomsae_syllabus": "taegeuk-3", "kyorugi_tier": "recreational",
        "primary_coach_full_name": "Dr Ali Alawi",
    }
    write_entity_sheet(wb, "Athletes", ATHLETE_COLUMNS, second_example=athlete_second)

    write_enums(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
